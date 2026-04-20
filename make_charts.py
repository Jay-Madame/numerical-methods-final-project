"""
make_charts.py — Generates results.xlsx from results.csv
One sheet per algorithm showing best, worst, and average time per matrix size,
ready for bar chart creation in Excel or Google Sheets.

Usage:
    python make_charts.py              # reads results.csv
    python make_charts.py my_data.csv  # custom CSV
"""

import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_FILE  = sys.argv[1] if len(sys.argv) > 1 else "results.csv"
OUT_FILE  = "results.xlsx"

# ── Colour palette per algorithm ──────────────────────────────────────────────
ALGO_STYLE = {
    "Jewell  (Hash + Pivot)": {"header_fill": "4472C4", "bar_fill": "9DC3E6"},
    "Carson  (LU Decomp)":    {"header_fill": "ED7D31", "bar_fill": "F4B183"},
    "Damian  (RREF)":         {"header_fill": "70AD47", "bar_fill": "A9D18E"},
    "Tyler   (Graph-based)":  {"header_fill": "7030A0", "bar_fill": "C5A3D4"},
}
DEFAULT_STYLE = {"header_fill": "595959", "bar_fill": "AEAAAA"}

SUMMARY_HEADER_FILL = "2E4057"


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def header_font(color="FFFFFF"):
    return Font(name="Arial", bold=True, color=color, size=11)


def body_font():
    return Font(name="Arial", size=10)


def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def center():
    return Alignment(horizontal="center", vertical="center")


def write_algo_sheet(wb, algo_name, df_algo):
    """Write one sheet for a single algorithm."""
    short = algo_name.split("(")[0].strip()
    ws = wb.create_sheet(title=short[:31])

    sizes = sorted(df_algo["matrix_size"].unique())
    style = ALGO_STYLE.get(algo_name, DEFAULT_STYLE)

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = f"{algo_name} — Best / Average / Worst Case (ms)"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = fill(style["header_fill"])
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 24

    # ── Column headers ─────────────────────────────────────────────────────
    headers = ["Matrix Size (n×n)", "Best Case (ms)", "Average Case (ms)",
               "Worst Case (ms)", "Std Dev (ms)"]
    col_widths = [22, 18, 20, 19, 16]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font()
        cell.fill = fill(style["header_fill"])
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[2].height = 18

    # ── Data rows ──────────────────────────────────────────────────────────
    data_start = 3
    for row_idx, n in enumerate(sizes, start=data_start):
        sub = df_algo[df_algo["matrix_size"] == n]["time_ms"]
        values = [
            f"{n}×{n}",
            round(float(sub.min()), 4),
            round(float(sub.mean()), 4),
            round(float(sub.max()), 4),
            round(float(sub.std(ddof=1)), 4),
        ]
        shade = "F2F2F2" if (row_idx - data_start) % 2 == 0 else "FFFFFF"
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = body_font()
            cell.fill = fill(shade)
            cell.alignment = center()
            cell.border = thin_border()

    # ── Named range label for charting hint ───────────────────────────────
    last_row = data_start + len(sizes) - 1
    note_row = last_row + 2
    ws.cell(row=note_row, column=1,
            value="↑ Select columns A–D above and insert a Clustered Bar Chart").font = \
        Font(name="Arial", italic=True, color="888888", size=9)

    return ws


def write_summary_sheet(wb, df):
    """Write a single comparison sheet with all algorithms side by side."""
    ws = wb.create_sheet(title="Summary (All Algorithms)", index=0)

    algos = sorted(df["algorithm"].unique())
    sizes = sorted(df["matrix_size"].unique())

    # Title
    total_cols = 1 + len(algos) * 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = "Gaussian Elimination — All Algorithms Comparison (ms)"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = fill(SUMMARY_HEADER_FILL)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 24

    # Sub-header: algorithm names spanning 3 cols each
    col = 2
    ws.cell(row=2, column=1, value="Matrix Size").font = header_font()
    ws.cell(row=2, column=1).fill = fill(SUMMARY_HEADER_FILL)
    ws.cell(row=2, column=1).alignment = center()
    ws.cell(row=2, column=1).border = thin_border()
    ws.column_dimensions["A"].width = 16

    algo_col_starts = {}
    for algo in algos:
        style = ALGO_STYLE.get(algo, DEFAULT_STYLE)
        short = algo.split("(")[0].strip()
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        cell = ws.cell(row=2, column=col, value=short)
        cell.font = header_font()
        cell.fill = fill(style["header_fill"])
        cell.alignment = center()
        cell.border = thin_border()
        algo_col_starts[algo] = col
        col += 3

    ws.row_dimensions[2].height = 18

    # Sub-sub-header: Best / Avg / Worst
    ws.cell(row=3, column=1, value="n×n").font = header_font()
    ws.cell(row=3, column=1).fill = fill(SUMMARY_HEADER_FILL)
    ws.cell(row=3, column=1).alignment = center()
    ws.cell(row=3, column=1).border = thin_border()

    for algo in algos:
        style = ALGO_STYLE.get(algo, DEFAULT_STYLE)
        c = algo_col_starts[algo]
        for offset, label in enumerate(["Best", "Avg", "Worst"]):
            cell = ws.cell(row=3, column=c + offset, value=label)
            cell.font = header_font()
            cell.fill = fill(style["header_fill"])
            cell.alignment = center()
            cell.border = thin_border()
            ws.column_dimensions[get_column_letter(c + offset)].width = 13

    ws.row_dimensions[3].height = 16

    # Data
    for row_idx, n in enumerate(sizes, start=4):
        shade = "F2F2F2" if row_idx % 2 == 0 else "FFFFFF"
        size_cell = ws.cell(row=row_idx, column=1, value=f"{n}×{n}")
        size_cell.font = body_font()
        size_cell.fill = fill(shade)
        size_cell.alignment = center()
        size_cell.border = thin_border()

        for algo in algos:
            sub = df[(df["algorithm"] == algo) & (df["matrix_size"] == n)]["time_ms"]
            c = algo_col_starts[algo]
            for offset, val in enumerate([
                round(float(sub.min()), 4) if len(sub) else "N/A",
                round(float(sub.mean()), 4) if len(sub) else "N/A",
                round(float(sub.max()), 4) if len(sub) else "N/A",
            ]):
                cell = ws.cell(row=row_idx, column=c + offset, value=val)
                cell.font = body_font()
                cell.fill = fill(shade)
                cell.alignment = center()
                cell.border = thin_border()

    last_row = 4 + len(sizes)
    note = ws.cell(row=last_row + 1, column=1,
                   value="↑ Select any row range + Best/Avg/Worst columns → Insert Clustered Bar Chart")
    note.font = Font(name="Arial", italic=True, color="888888", size=9)

    return ws


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"\n  Reading '{CSV_FILE}'...")
    df = pd.read_csv(CSV_FILE)
    df = df[df["status"] == "OK"].copy()
    df["time_ms"] = pd.to_numeric(df["time_ms"], errors="coerce")
    df = df.dropna(subset=["time_ms"])
    df["algorithm"] = df["algorithm"].str.strip()

    algos = sorted(df["algorithm"].unique())
    print(f"  Found {len(algos)} algorithms: {algos}")
    print(f"  Matrix sizes: {sorted(df['matrix_size'].unique())}")

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    write_summary_sheet(wb, df)

    for algo in algos:
        df_algo = df[df["algorithm"] == algo]
        write_algo_sheet(wb, algo, df_algo)
        print(f"  ✓ Sheet written: {algo.split('(')[0].strip()}")

    wb.save(OUT_FILE)
    print(f"\n  Saved → '{OUT_FILE}'")
    print(f"  Sheets: Summary + one per algorithm ({len(algos)} total)\n")


if __name__ == "__main__":
    main()
